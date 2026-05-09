# 海关数据自动分析脚本（带图表版）
# 使用前先在终端运行：pip install pandas openpyxl openai matplotlib

import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from datetime import datetime
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import io
import os

# ============================================================
# 修改这里的配置
# ============================================================
EXCEL_FILE  = r"D:\software\Microsoft VS Code\尼日利亚海关数据.xlsx"
API_KEY     = "your_api_key_here"
OUTPUT_FILE = f"分析报告_{datetime.now().strftime('%Y%m%d')}.xlsx"
# ============================================================

# 解决 matplotlib 中文乱码
# 自动寻找系统中文字体
def set_chinese_font():
    candidates = ["Microsoft YaHei", "SimHei", "STHeiti", "PingFang SC", "Arial Unicode MS"]
    for font in candidates:
        if any(font.lower() in f.name.lower() for f in fm.fontManager.ttflist):
            matplotlib.rcParams["font.family"] = font
            matplotlib.rcParams["axes.unicode_minus"] = False
            return
    # 找不到就用系统默认，标签改英文兜底
    matplotlib.rcParams["axes.unicode_minus"] = False

set_chinese_font()

# 连接千问 API
client = OpenAI(
    api_key=API_KEY,
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
)

# ──────────────────────────────────────────
# 数据读取（自动合并所有 Sheet）
# ──────────────────────────────────────────
def load_data(filepath):
    print(f"正在读取：{filepath}")
    all_sheets = pd.read_excel(filepath, sheet_name=None)  # 读取全部 Sheet
    print(f"发现 {len(all_sheets)} 个 Sheet：{', '.join(all_sheets.keys())}")

    frames = []
    for sheet_name, sheet_df in all_sheets.items():
        sheet_df = sheet_df.copy()
        sheet_df["来源国家Sheet"] = sheet_name   # 记录来自哪个 Sheet
        frames.append(sheet_df)

    df = pd.concat(frames, ignore_index=True)
    df["日期"] = pd.to_datetime(df["日期"], errors="coerce")
    df["金额"] = pd.to_numeric(df["金额"], errors="coerce")
    df["数量"] = pd.to_numeric(df["数量"], errors="coerce")
    df["重量"] = pd.to_numeric(df["重量"], errors="coerce")
    print(f"✅ 合并后共 {len(df)} 条记录，时间范围：{df['日期'].min().date()} ~ {df['日期'].max().date()}")
    return df

# ──────────────────────────────────────────
# 数据分析
# ──────────────────────────────────────────
def analyze_buyers(df, top_n=20):
    buyers = (
        df.groupby("采购商")
        .agg(
            交易次数=("金额", "count"),
            总金额=("金额", "sum"),
            总数量=("数量", "sum"),
            涉及产品=("产品描述", lambda x: "、".join(x.dropna().unique()[:3])),
            进口国家=("进口国/地区", lambda x: "、".join(x.dropna().unique()[:3])),
        )
        .sort_values("总金额", ascending=False)
        .head(top_n).reset_index()
    )
    buyers["总金额"] = buyers["总金额"].round(2)
    return buyers

def analyze_competitors(df, top_n=20):
    suppliers = (
        df.groupby(["供应商", "销售国/地区"])
        .agg(
            交易次数=("金额", "count"),
            总金额=("金额", "sum"),
            主要产品=("产品描述", lambda x: "、".join(x.dropna().unique()[:3])),
            主要客户=("采购商", lambda x: "、".join(x.dropna().unique()[:3])),
        )
        .sort_values("总金额", ascending=False)
        .head(top_n).reset_index()
    )
    suppliers["总金额"] = suppliers["总金额"].round(2)
    return suppliers

def analyze_trends(df):
    df = df.copy()
    df["月份"] = df["日期"].dt.to_period("M").astype(str)
    trends = (
        df.groupby("月份")
        .agg(
            交易次数=("金额", "count"),
            总金额=("金额", "sum"),
            总数量=("数量", "sum"),
            活跃采购商=("采购商", "nunique"),
            活跃供应商=("供应商", "nunique"),
        )
        .sort_index().reset_index()
    )
    trends["总金额"] = trends["总金额"].round(2)
    return trends

def analyze_hs_codes(df, top_n=15):
    hs = (
        df.groupby(["HS编码", "产品描述"])
        .agg(
            交易次数=("金额", "count"),
            总金额=("金额", "sum"),
            采购商数量=("采购商", "nunique"),
        )
        .sort_values("总金额", ascending=False)
        .head(top_n).reset_index()
    )
    hs["总金额"] = hs["总金额"].round(2)
    return hs

# ──────────────────────────────────────────
# 图表生成（返回图片字节流）
# ──────────────────────────────────────────
def chart_to_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf

def make_trend_chart(trends):
    """月度金额折线图"""
    fig, ax1 = plt.subplots(figsize=(12, 5))
    x = range(len(trends))
    ax1.plot(x, trends["总金额"], color="#2563EB", linewidth=2.5, marker="o", markersize=5, label="总金额")
    ax1.fill_between(x, trends["总金额"], alpha=0.1, color="#2563EB")
    ax1.set_ylabel("总金额", color="#2563EB")
    ax1.tick_params(axis="y", labelcolor="#2563EB")
    ax2 = ax1.twinx()
    ax2.bar(x, trends["交易次数"], alpha=0.3, color="#10B981", label="交易次数")
    ax2.set_ylabel("交易次数", color="#10B981")
    ax2.tick_params(axis="y", labelcolor="#10B981")
    ax1.set_xticks(list(x))
    ax1.set_xticklabels(trends["月份"].tolist(), rotation=45, ha="right", fontsize=8)
    ax1.set_title("月度贸易趋势（金额 + 交易次数）", fontsize=13, pad=12)
    fig.tight_layout()
    return chart_to_bytes(fig)

def make_buyers_chart(buyers):
    """Top10 采购商横向柱状图"""
    top = buyers.head(10).iloc[::-1]
    # 截断过长名称
    labels = [str(n)[:20] + "…" if len(str(n)) > 20 else str(n) for n in top["采购商"]]
    fig, ax = plt.subplots(figsize=(11, 6))
    bars = ax.barh(labels, top["总金额"], color="#3B82F6", edgecolor="white", height=0.6)
    for bar, val in zip(bars, top["总金额"]):
        ax.text(bar.get_width() * 1.01, bar.get_y() + bar.get_height() / 2,
                f"{val:,.0f}", va="center", fontsize=8)
    ax.set_xlabel("总金额")
    ax.set_title("Top 10 采购商（按总金额）", fontsize=13, pad=12)
    ax.spines[["top", "right"]].set_visible(False)
    fig.tight_layout()
    return chart_to_bytes(fig)

def make_suppliers_chart(competitors):
    """Top10 供应商横向柱状图"""
    top = competitors.head(10).iloc[::-1]
    labels = [str(n)[:20] + "…" if len(str(n)) > 20 else str(n) for n in top["供应商"]]
    fig, ax = plt.subplots(figsize=(11, 6))
    bars = ax.barh(labels, top["总金额"], color="#8B5CF6", edgecolor="white", height=0.6)
    for bar, val in zip(bars, top["总金额"]):
        ax.text(bar.get_width() * 1.01, bar.get_y() + bar.get_height() / 2,
                f"{val:,.0f}", va="center", fontsize=8)
    ax.set_xlabel("总金额")
    ax.set_title("Top 10 供应商/竞争对手（按总金额）", fontsize=13, pad=12)
    ax.spines[["top", "right"]].set_visible(False)
    fig.tight_layout()
    return chart_to_bytes(fig)

def make_hs_chart(hs_codes):
    """Top10 HS 编码饼图"""
    top = hs_codes.head(10)
    labels = [f"{str(row['HS编码'])} {str(row['产品描述'])[:12]}" for _, row in top.iterrows()]
    fig, ax = plt.subplots(figsize=(10, 7))
    wedges, texts, autotexts = ax.pie(
        top["总金额"],
        labels=None,
        autopct="%1.1f%%",
        startangle=140,
        colors=plt.cm.Set3.colors[:len(top)],
        pctdistance=0.8,
    )
    ax.legend(wedges, labels, title="HS编码 产品", loc="center left",
              bbox_to_anchor=(1, 0, 0.5, 1), fontsize=8)
    ax.set_title("Top 10 HS编码 金额占比", fontsize=13, pad=12)
    fig.tight_layout()
    return chart_to_bytes(fig)

def make_country_chart(df):
    """进口国家/地区分布柱状图"""
    country = (
        df.groupby("进口国/地区")["金额"].sum()
        .sort_values(ascending=False)
        .head(10)
        .iloc[::-1]
    )
    fig, ax = plt.subplots(figsize=(11, 5))
    bars = ax.barh(country.index.tolist(), country.values, color="#F59E0B", edgecolor="white", height=0.6)
    for bar, val in zip(bars, country.values):
        ax.text(bar.get_width() * 1.01, bar.get_y() + bar.get_height() / 2,
                f"{val:,.0f}", va="center", fontsize=8)
    ax.set_xlabel("总金额")
    ax.set_title("Top 10 进口国家/地区（按总金额）", fontsize=13, pad=12)
    ax.spines[["top", "right"]].set_visible(False)
    fig.tight_layout()
    return chart_to_bytes(fig)

def make_sheet_compare_chart(df):
    """各国家 Sheet 对比图（金额 + 交易次数 + 采购商数）"""
    sheet_stats = (
        df.groupby("来源国家Sheet")
        .agg(总金额=("金额", "sum"), 交易次数=("金额", "count"), 采购商数=("采购商", "nunique"))
        .sort_values("总金额", ascending=False)
        .reset_index()
    )
    n = len(sheet_stats)
    x = range(n)
    labels = sheet_stats["来源国家Sheet"].tolist()

    fig, axes = plt.subplots(1, 3, figsize=(14, 5))
    colors = plt.cm.Set2.colors

    # 金额柱状图
    axes[0].bar(x, sheet_stats["总金额"], color=colors[:n], edgecolor="white")
    axes[0].set_xticks(list(x)); axes[0].set_xticklabels(labels, rotation=30, ha="right", fontsize=8)
    axes[0].set_title("各Sheet 总金额", fontsize=11)
    axes[0].spines[["top", "right"]].set_visible(False)
    for i, v in enumerate(sheet_stats["总金额"]):
        axes[0].text(i, v * 1.01, f"{v:,.0f}", ha="center", fontsize=7)

    # 交易次数
    axes[1].bar(x, sheet_stats["交易次数"], color=colors[:n], edgecolor="white")
    axes[1].set_xticks(list(x)); axes[1].set_xticklabels(labels, rotation=30, ha="right", fontsize=8)
    axes[1].set_title("各Sheet 交易次数", fontsize=11)
    axes[1].spines[["top", "right"]].set_visible(False)
    for i, v in enumerate(sheet_stats["交易次数"]):
        axes[1].text(i, v * 1.01, str(v), ha="center", fontsize=7)

    # 采购商数量
    axes[2].bar(x, sheet_stats["采购商数"], color=colors[:n], edgecolor="white")
    axes[2].set_xticks(list(x)); axes[2].set_xticklabels(labels, rotation=30, ha="right", fontsize=8)
    axes[2].set_title("各Sheet 采购商数量", fontsize=11)
    axes[2].spines[["top", "right"]].set_visible(False)
    for i, v in enumerate(sheet_stats["采购商数"]):
        axes[2].text(i, v * 1.01, str(v), ha="center", fontsize=7)

    fig.suptitle("各国家市场对比总览", fontsize=13, y=1.02)
    fig.tight_layout()
    return chart_to_bytes(fig)

def analyze_by_sheet(df):
    """按国家 Sheet 汇总统计表"""
    stats = (
        df.groupby("来源国家Sheet")
        .agg(
            总金额=("金额", "sum"),
            交易次数=("金额", "count"),
            采购商数量=("采购商", "nunique"),
            供应商数量=("供应商", "nunique"),
            平均单笔金额=("金额", "mean"),
            最大单笔金额=("金额", "max"),
        )
        .sort_values("总金额", ascending=False)
        .reset_index()
    )
    stats["总金额"]     = stats["总金额"].round(2)
    stats["平均单笔金额"] = stats["平均单笔金额"].round(2)
    stats["最大单笔金额"] = stats["最大单笔金额"].round(2)
    return stats

# ──────────────────────────────────────────
# AI 分析
# ──────────────────────────────────────────
def build_summary(df, buyers, competitors, trends, hs_codes):
    lines = []
    lines.append(f"数据概况：共{len(df)}条记录，时间范围 {df['日期'].min().date()} 至 {df['日期'].max().date()}")
    lines.append(f"总贸易金额：{df['金额'].sum():,.0f}，涉及采购商 {df['采购商'].nunique()} 家，供应商 {df['供应商'].nunique()} 家")
    lines.append("\n【TOP5 采购商】")
    for _, r in buyers.head(5).iterrows():
        lines.append(f"  {r['采购商']}：金额 {r['总金额']:,.0f}，交易 {r['交易次数']} 次，产品：{r['涉及产品']}")
    lines.append("\n【TOP5 供应商（竞争对手）】")
    for _, r in competitors.head(5).iterrows():
        lines.append(f"  {r['供应商']}（{r['销售国/地区']}）：金额 {r['总金额']:,.0f}，客户：{r['主要客户']}")
    lines.append("\n【近6个月趋势】")
    for _, r in trends.tail(6).iterrows():
        lines.append(f"  {r['月份']}：金额 {r['总金额']:,.0f}，{r['交易次数']} 笔，{r['活跃采购商']} 家采购商")
    lines.append("\n【TOP5 HS编码】")
    for _, r in hs_codes.head(5).iterrows():
        lines.append(f"  {r['HS编码']} {r['产品描述']}：金额 {r['总金额']:,.0f}，{r['采购商数量']} 家客户")
    return "\n".join(lines)

def ask_ai(summary_text):
    print("正在请求 AI 分析，请稍候（约10-20秒）...")
    prompt = f"""
你是一位外贸行业分析师，请根据以下海关数据汇总，给出专业的分析报告。

{summary_text}

请从以下四个角度分析，每点100字左右：
1. 【潜在客户机会】哪些采购商值得重点开发，理由是什么
2. 【竞争对手洞察】主要竞争供应商的特点和策略
3. 【市场趋势判断】市场整体走势和季节性规律
4. 【业务建议】基于数据给出3条具体可执行的建议

用中文回答，语言专业简洁。
"""
    response = client.chat.completions.create(
        model="qwen-plus",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=1500,
    )
    return response.choices[0].message.content

# ──────────────────────────────────────────
# 保存报告（数据表 + 图表）
# ──────────────────────────────────────────
def save_report(df, buyers, competitors, trends, hs_codes, sheet_stats, ai_conclusion, charts, output_file):
    print(f"正在生成报告：{output_file}")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        pd.DataFrame({"AI分析结论": [ai_conclusion]}).to_excel(writer, sheet_name="AI分析结论",   index=False)
        sheet_stats.to_excel(writer,  sheet_name="各国家市场对比",  index=False)
        buyers.to_excel(writer,       sheet_name="潜在客户排名",    index=False)
        competitors.to_excel(writer,  sheet_name="竞争对手排名",    index=False)
        trends.to_excel(writer,       sheet_name="月度趋势",        index=False)
        hs_codes.to_excel(writer,     sheet_name="HS编码分析",      index=False)
        writer.book.create_sheet("📊 图表总览")

    wb = load_workbook(output_file)
    ws = wb["📊 图表总览"]
    chart_configs = [
        (charts["sheet_compare"], "A1"),
        (charts["trend"],         "A32"),
        (charts["buyers"],        "A62"),
        (charts["suppliers"],     "A92"),
        (charts["hs"],            "A122"),
        (charts["country"],       "A152"),
    ]
    for img_buf, cell in chart_configs:
        img = XLImage(img_buf)
        img.width  = 750
        img.height = 350
        ws.add_image(img, cell)

    wb.save(output_file)
    print(f"✅ 报告已生成：{output_file}")

# ──────────────────────────────────────────
# 主流程
# ──────────────────────────────────────────
def main():
    print("=" * 50)
    print("海关数据自动分析工具（图表版）")
    print("=" * 50)

    df          = load_data(EXCEL_FILE)

    print("正在分析采购商...")
    buyers      = analyze_buyers(df)

    print("正在分析供应商...")
    competitors = analyze_competitors(df)

    print("正在分析月度趋势...")
    trends      = analyze_trends(df)

    print("正在分析 HS 编码...")
    hs_codes    = analyze_hs_codes(df)

    print("正在分析各国家Sheet...")
    sheet_stats = analyze_by_sheet(df)

    print("正在生成图表...")
    charts = {
        "sheet_compare": make_sheet_compare_chart(df),
        "trend":         make_trend_chart(trends),
        "buyers":        make_buyers_chart(buyers),
        "suppliers":     make_suppliers_chart(competitors),
        "hs":            make_hs_chart(hs_codes),
        "country":       make_country_chart(df),
    }
    print("✅ 图表生成完成")

    summary      = build_summary(df, buyers, competitors, trends, hs_codes)
    ai_conclusion = ask_ai(summary)
    print("\n【AI 分析结论】")
    print(ai_conclusion)

    save_report(df, buyers, competitors, trends, hs_codes, sheet_stats, ai_conclusion, charts, OUTPUT_FILE)
    print("\n🎉 全部完成！报告保存在：", os.path.abspath(OUTPUT_FILE))

if __name__ == "__main__":
    main()
