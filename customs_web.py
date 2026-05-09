# 海关数据分析系统 —— 网页版
# 启动方法：在终端输入  streamlit run customs_web.py

import streamlit as st
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import io
from datetime import datetime
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ──────────────────────────────────────────
# 页面基础配置
# ──────────────────────────────────────────
st.set_page_config(
    page_title="海关数据分析系统",
    page_icon="🌍",
    layout="wide",
)

# ──────────────────────────────────────────
# 中文字体
# ──────────────────────────────────────────
def set_chinese_font():
    candidates = ["Microsoft YaHei", "SimHei", "STHeiti", "PingFang SC", "Arial Unicode MS"]
    for font in candidates:
        if any(font.lower() in f.name.lower() for f in fm.fontManager.ttflist):
            matplotlib.rcParams["font.family"] = font
            matplotlib.rcParams["axes.unicode_minus"] = False
            return
    matplotlib.rcParams["axes.unicode_minus"] = False

set_chinese_font()

# ──────────────────────────────────────────
# 侧边栏：配置区
# ──────────────────────────────────────────
with st.sidebar:
    st.image("https://img.icons8.com/color/96/trade.png", width=60)
    st.title("海关数据分析系统")
    st.markdown("---")
    st.subheader("⚙️ 配置")
    api_key = st.text_input(
        "千问 API Key",
        type="password",
        placeholder="sk-xxxxxxxxxxxxxxxx",
        help="填入你的阿里云千问 API Key"
    )
    st.markdown("---")
    st.markdown("**使用步骤**")
    st.markdown("1. 填入 API Key\n2. 上传 Excel 文件\n3. 点击「开始分析」\n4. 查看报告 / 下载")
    st.markdown("---")
    st.caption("支持多 Sheet，自动按国家合并")

# ──────────────────────────────────────────
# 主页面标题
# ──────────────────────────────────────────
st.markdown("## 🌍 海关数据智能分析系统")
st.markdown("上传海关原始数据 Excel，自动生成多维度分析报告，支持 AI 洞察与一键下载。")
st.markdown("---")

# ──────────────────────────────────────────
# 文件上传区
# ──────────────────────────────────────────
uploaded_file = st.file_uploader(
    "📂 上传海关数据 Excel 文件（支持多 Sheet）",
    type=["xlsx", "xls"],
    help="支持按国家分 Sheet 的格式，自动合并分析"
)

# ──────────────────────────────────────────
# 数据处理函数
# ──────────────────────────────────────────
@st.cache_data
def load_data(file_bytes):
    all_sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)
    frames = []
    for sheet_name, sheet_df in all_sheets.items():
        sheet_df = sheet_df.copy()
        sheet_df["来源国家Sheet"] = sheet_name
        frames.append(sheet_df)
    df = pd.concat(frames, ignore_index=True)
    df["日期"] = pd.to_datetime(df["日期"], errors="coerce")
    df["金额"] = pd.to_numeric(df["金额"], errors="coerce")
    df["数量"] = pd.to_numeric(df["数量"], errors="coerce")
    df["重量"] = pd.to_numeric(df["重量"], errors="coerce")
    return df, list(all_sheets.keys())

def analyze_buyers(df, top_n=20):
    return (
        df.groupby("采购商").agg(
            交易次数=("金额", "count"),
            总金额=("金额", "sum"),
            总数量=("数量", "sum"),
            涉及产品=("产品描述", lambda x: "、".join(x.dropna().unique()[:3])),
            进口国家=("进口国/地区", lambda x: "、".join(x.dropna().unique()[:3])),
        ).sort_values("总金额", ascending=False).head(top_n).reset_index()
    )

def analyze_competitors(df, top_n=20):
    return (
        df.groupby(["供应商", "销售国/地区"]).agg(
            交易次数=("金额", "count"),
            总金额=("金额", "sum"),
            主要产品=("产品描述", lambda x: "、".join(x.dropna().unique()[:3])),
            主要客户=("采购商", lambda x: "、".join(x.dropna().unique()[:3])),
        ).sort_values("总金额", ascending=False).head(top_n).reset_index()
    )

def analyze_trends(df):
    d = df.copy()
    d["月份"] = d["日期"].dt.to_period("M").astype(str)
    return (
        d.groupby("月份").agg(
            交易次数=("金额", "count"),
            总金额=("金额", "sum"),
            活跃采购商=("采购商", "nunique"),
            活跃供应商=("供应商", "nunique"),
        ).sort_index().reset_index()
    )

def analyze_hs_codes(df, top_n=15):
    return (
        df.groupby(["HS编码", "产品描述"]).agg(
            交易次数=("金额", "count"),
            总金额=("金额", "sum"),
            采购商数量=("采购商", "nunique"),
        ).sort_values("总金额", ascending=False).head(top_n).reset_index()
    )

def analyze_by_sheet(df):
    stats = (
        df.groupby("来源国家Sheet").agg(
            总金额=("金额", "sum"),
            交易次数=("金额", "count"),
            采购商数量=("采购商", "nunique"),
            供应商数量=("供应商", "nunique"),
            平均单笔金额=("金额", "mean"),
            最大单笔金额=("金额", "max"),
        ).sort_values("总金额", ascending=False).reset_index()
    )
    for col in ["总金额", "平均单笔金额", "最大单笔金额"]:
        stats[col] = stats[col].round(2)
    return stats

# ──────────────────────────────────────────
# 图表函数（返回 figure，Streamlit 直接渲染）
# ──────────────────────────────────────────
def fig_trend(trends):
    fig, ax1 = plt.subplots(figsize=(12, 4))
    x = range(len(trends))
    ax1.plot(x, trends["总金额"], color="#2563EB", linewidth=2.5, marker="o", markersize=5)
    ax1.fill_between(x, trends["总金额"], alpha=0.1, color="#2563EB")
    ax1.set_ylabel("总金额", color="#2563EB")
    ax1.tick_params(axis="y", labelcolor="#2563EB")
    ax2 = ax1.twinx()
    ax2.bar(x, trends["交易次数"], alpha=0.3, color="#10B981")
    ax2.set_ylabel("交易次数", color="#10B981")
    ax1.set_xticks(list(x))
    ax1.set_xticklabels(trends["月份"].tolist(), rotation=45, ha="right", fontsize=8)
    ax1.set_title("月度贸易趋势", fontsize=12)
    fig.tight_layout()
    return fig

def fig_barh(series_name, series_values, title, color):
    vals = series_values.iloc[::-1]
    names = series_name.iloc[::-1]
    labels = [str(n)[:22] + "…" if len(str(n)) > 22 else str(n) for n in names]
    fig, ax = plt.subplots(figsize=(10, 5))
    bars = ax.barh(labels, vals, color=color, edgecolor="white", height=0.6)
    for bar, val in zip(bars, vals):
        ax.text(bar.get_width() * 1.01, bar.get_y() + bar.get_height() / 2,
                f"{val:,.0f}", va="center", fontsize=8)
    ax.set_title(title, fontsize=12)
    ax.spines[["top", "right"]].set_visible(False)
    fig.tight_layout()
    return fig

def fig_pie(hs_codes):
    top = hs_codes.head(10)
    labels = [f"{str(r['HS编码'])} {str(r['产品描述'])[:10]}" for _, r in top.iterrows()]
    fig, ax = plt.subplots(figsize=(9, 6))
    wedges, _, autotexts = ax.pie(
        top["总金额"], labels=None, autopct="%1.1f%%",
        startangle=140, colors=plt.cm.Set3.colors[:len(top)], pctdistance=0.8,
    )
    ax.legend(wedges, labels, title="HS编码", loc="center left",
              bbox_to_anchor=(1, 0, 0.5, 1), fontsize=8)
    ax.set_title("HS编码金额占比", fontsize=12)
    fig.tight_layout()
    return fig

def fig_sheet_compare(df):
    s = df.groupby("来源国家Sheet").agg(
        总金额=("金额","sum"), 交易次数=("金额","count"), 采购商数=("采购商","nunique")
    ).sort_values("总金额", ascending=False).reset_index()
    n = len(s); x = range(n)
    labels = s["来源国家Sheet"].tolist()
    colors = plt.cm.Set2.colors
    fig, axes = plt.subplots(1, 3, figsize=(13, 4))
    for ax, col, ttl in zip(axes, ["总金额","交易次数","采购商数"], ["各市场总金额","各市场交易次数","各市场采购商数"]):
        ax.bar(x, s[col], color=colors[:n], edgecolor="white")
        ax.set_xticks(list(x)); ax.set_xticklabels(labels, rotation=25, ha="right", fontsize=8)
        ax.set_title(ttl, fontsize=10)
        ax.spines[["top","right"]].set_visible(False)
        for i, v in enumerate(s[col]):
            ax.text(i, v*1.01, f"{v:,.0f}" if col=="总金额" else str(v), ha="center", fontsize=7)
    fig.suptitle("各国家市场横向对比", fontsize=12)
    fig.tight_layout()
    return fig

# 把 matplotlib figure 转成字节流（用于写入 Excel）
def fig_to_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    return buf

# ──────────────────────────────────────────
# AI 分析
# ──────────────────────────────────────────
def ask_ai(api_key, df, buyers, competitors, trends, hs_codes):
    client = OpenAI(api_key=api_key, base_url="https://dashscope.aliyuncs.com/compatible-mode/v1")
    lines = [
        f"数据概况：共{len(df)}条记录，时间范围 {df['日期'].min().date()} 至 {df['日期'].max().date()}",
        f"总金额：{df['金额'].sum():,.0f}，采购商 {df['采购商'].nunique()} 家，供应商 {df['供应商'].nunique()} 家",
        "\n【TOP5 采购商】",
    ]
    for _, r in buyers.head(5).iterrows():
        lines.append(f"  {r['采购商']}：金额 {r['总金额']:,.0f}，{r['交易次数']} 次，产品：{r['涉及产品']}")
    lines.append("\n【TOP5 供应商】")
    for _, r in competitors.head(5).iterrows():
        lines.append(f"  {r['供应商']}（{r['销售国/地区']}）：金额 {r['总金额']:,.0f}")
    lines.append("\n【近6个月趋势】")
    for _, r in trends.tail(6).iterrows():
        lines.append(f"  {r['月份']}：金额 {r['总金额']:,.0f}，{r['交易次数']} 笔")
    lines.append("\n【TOP5 HS编码】")
    for _, r in hs_codes.head(5).iterrows():
        lines.append(f"  {r['HS编码']} {r['产品描述']}：金额 {r['总金额']:,.0f}")
    summary = "\n".join(lines)

    prompt = f"""你是一位外贸行业分析师，请根据以下海关数据给出专业分析。

{summary}

请从四个角度分析，每点约100字：
1. 【潜在客户机会】值得重点开发的采购商及理由
2. 【竞争对手洞察】主要供应商的特点和策略
3. 【市场趋势判断】整体走势和季节性规律
4. 【业务建议】3条具体可执行的建议

用中文回答，语言专业简洁。"""
    resp = client.chat.completions.create(
        model="qwen-plus",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=1500,
    )
    return resp.choices[0].message.content

# ──────────────────────────────────────────
# 生成可下载的 Excel 报告
# ──────────────────────────────────────────
def build_excel(buyers, competitors, trends, hs_codes, sheet_stats, ai_text, figs):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame({"AI分析结论": [ai_text]}).to_excel(writer, sheet_name="AI分析结论", index=False)
        sheet_stats.to_excel(writer,  sheet_name="各国家市场对比", index=False)
        buyers.to_excel(writer,       sheet_name="潜在客户排名",   index=False)
        competitors.to_excel(writer,  sheet_name="竞争对手排名",   index=False)
        trends.to_excel(writer,       sheet_name="月度趋势",       index=False)
        hs_codes.to_excel(writer,     sheet_name="HS编码分析",     index=False)
        writer.book.create_sheet("图表总览")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb["图表总览"]
    cells = ["A1", "A32", "A62", "A92", "A122", "A152"]
    for fig, cell in zip(figs, cells):
        buf = fig_to_bytes(fig)
        img = XLImage(buf)
        img.width = 720; img.height = 340
        ws.add_image(img, cell)

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final

# ──────────────────────────────────────────
# 主流程：上传后触发
# ──────────────────────────────────────────
if uploaded_file:
    file_bytes = uploaded_file.read()
    df, sheet_names = load_data(file_bytes)

    # 数据概览卡片
    st.markdown("### 📋 数据概览")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("总记录数",   f"{len(df):,}")
    c2.metric("总交易金额", f"{df['金额'].sum():,.0f}")
    c3.metric("采购商数量", f"{df['采购商'].nunique()}")
    c4.metric("供应商数量", f"{df['供应商'].nunique()}")
    c5.metric("覆盖国家数", f"{len(sheet_names)}")

    st.caption(f"数据时间范围：{df['日期'].min().date()} ~ {df['日期'].max().date()}　｜　识别到 {len(sheet_names)} 个 Sheet：{' / '.join(sheet_names)}")
    st.markdown("---")

    # 开始分析按钮
    if st.button("🚀 开始分析", type="primary", use_container_width=True):
        if not api_key:
            st.error("请先在左侧填入千问 API Key")
            st.stop()

        with st.spinner("正在分析数据，请稍候..."):
            buyers      = analyze_buyers(df)
            competitors = analyze_competitors(df)
            trends      = analyze_trends(df)
            hs_codes    = analyze_hs_codes(df)
            sheet_stats = analyze_by_sheet(df)

        with st.spinner("正在请求 AI 分析（约15秒）..."):
            ai_text = ask_ai(api_key, df, buyers, competitors, trends, hs_codes)

        st.success("✅ 分析完成！")
        st.markdown("---")

        # ── AI 分析结论 ──
        st.markdown("### 🤖 AI 分析结论")
        st.info(ai_text)
        st.markdown("---")

        # ── 各国家市场对比 ──
        st.markdown("### 🌐 各国家市场对比")
        f_sheet = fig_sheet_compare(df)
        st.pyplot(f_sheet)
        st.dataframe(sheet_stats, use_container_width=True, hide_index=True)
        st.markdown("---")

        # ── 月度趋势 ──
        st.markdown("### 📈 月度贸易趋势")
        f_trend = fig_trend(trends)
        st.pyplot(f_trend)
        st.dataframe(trends, use_container_width=True, hide_index=True)
        st.markdown("---")

        # ── 采购商 & 供应商并排 ──
        st.markdown("### 👥 采购商 & 供应商分析")
        col_l, col_r = st.columns(2)
        with col_l:
            st.markdown("**Top 10 潜在客户（采购商）**")
            f_buyers = fig_barh(buyers["采购商"].head(10), buyers["总金额"].head(10), "Top 10 采购商", "#3B82F6")
            st.pyplot(f_buyers)
        with col_r:
            st.markdown("**Top 10 竞争对手（供应商）**")
            f_supp = fig_barh(competitors["供应商"].head(10), competitors["总金额"].head(10), "Top 10 供应商", "#8B5CF6")
            st.pyplot(f_supp)

        st.markdown("**采购商详情**")
        st.dataframe(buyers, use_container_width=True, hide_index=True)
        st.markdown("**供应商详情**")
        st.dataframe(competitors, use_container_width=True, hide_index=True)
        st.markdown("---")

        # ── HS 编码 ──
        st.markdown("### 📦 HS 编码 / 产品结构")
        f_hs = fig_pie(hs_codes)
        st.pyplot(f_hs)
        st.dataframe(hs_codes, use_container_width=True, hide_index=True)
        st.markdown("---")

        # ── 下载报告 ──
        st.markdown("### 💾 下载分析报告")
        figs_for_excel = [f_sheet, f_trend, f_buyers, f_supp, f_hs]
        excel_bytes = build_excel(buyers, competitors, trends, hs_codes, sheet_stats, ai_text, figs_for_excel)
        filename = f"海关分析报告_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            label="📥 下载 Excel 报告（含图表）",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

else:
    # 未上传时的提示界面
    st.markdown("### 👆 请上传海关数据文件开始分析")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("**📊 支持多维度分析**")
        st.markdown("采购商排名 / 供应商对比 / 月度趋势 / HS编码分布 / 国家市场对比")
    with col2:
        st.markdown("**🤖 AI 智能洞察**")
        st.markdown("自动生成客户机会、竞争分析、趋势判断、业务建议")
    with col3:
        st.markdown("**💾 一键下载报告**")
        st.markdown("含图表的完整 Excel 报告，可直接用于汇报展示")
